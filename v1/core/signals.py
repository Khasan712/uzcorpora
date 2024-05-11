from django.db.models.signals import pre_save
from django.dispatch import receiver
from v1.core.models import CreateWordFromExcel, Word, WordGrammar, WordSemanticExpression
from openpyxl import load_workbook
from v1.utils.validations import validate_word_apostrophe
from django.core.validators import ValidationError
from django.db import transaction
import uuid


# @receiver(post_save, sender=Text)
# def text_signals(sender, instance, created, **kwargs):
#     if created and getattr(instance, 'id', None):
#         text_validate_and_config_task.delay(instance.id)


@receiver(pre_save, sender=CreateWordFromExcel)
def create_word_model_signal(sender, instance, **kwargs):
    try:
        if instance.id is None:
            uuid_code = uuid.uuid4()
            instance.uuid = uuid_code
            with transaction.atomic():
                wb = load_workbook(instance.file)
                sheet = wb.active
                items_ids = list(sheet.iter_rows(values_only=True))[0][2:]
                word_phrase = instance.word_phrase

                if word_phrase == 'grammar':
                    for row in list(sheet.iter_rows(values_only=True))[1:]:
                        word = validate_word_apostrophe(str(row[0]))
                        lemma = validate_word_apostrophe(str(row[1]))

                        word_obj = Word.objects.create(
                            phrase=instance.phrase, word=word, lemma=lemma, uuid=uuid_code
                        )
                        semantics = row[2:]
                        semantics_obj = []
                        for i, semantic in enumerate(semantics):
                            if semantic and str(semantic).strip() == '+':
                                semantics_obj.append(
                                    WordGrammar(word=word_obj, phrase_id=items_ids[i])
                                )
                        if semantics_obj:
                            WordGrammar.objects.bulk_create(semantics_obj)

                elif word_phrase == 'semantic_expression':
                    for row in list(sheet.iter_rows(values_only=True))[1:]:
                        word = validate_word_apostrophe(str(row[0]))
                        lemma = validate_word_apostrophe(str(row[1]))

                        word_obj = Word.objects.create(
                            phrase=instance.phrase, word=word, lemma=lemma, uuid=uuid_code
                        )
                        semantics = row[2:]
                        semantics_obj = []
                        for i, semantic in enumerate(semantics):
                            if semantic and str(semantic).strip() == '+':
                                semantics_obj.append(
                                    WordSemanticExpression(word=word_obj, phrase_id=items_ids[i])
                                )
                        if semantics_obj:
                            WordSemanticExpression.objects.bulk_create(semantics_obj)
    except Exception as e:
        raise ValidationError("Error  => " + str(e))

